from __future__ import annotations

import io
import uuid

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.poll import Option
from app.models.user import AdminRole, AdminUser
from app.services.auth import create_access_token, hash_password, verify_password

pytestmark = pytest.mark.asyncio


async def _create_admin_and_csrf(client: AsyncClient, db: AsyncSession) -> tuple[AdminUser, str]:
    admin = AdminUser(
        username=f"admin_{uuid.uuid4().hex[:8]}",
        password_hash=hash_password("admin_password_123"),
        role=AdminRole.ADMIN,
        is_active=True,
    )
    db.add(admin)
    await db.commit()

    client.cookies.set("access_token", create_access_token(admin.id))
    response = await client.get("/setup")
    csrf_token = response.cookies.get("csrftoken")
    assert csrf_token is not None
    return admin, csrf_token


async def test_users_crud_and_permission_checks(client: AsyncClient, db_session: AsyncSession) -> None:
    admin, csrf_token = await _create_admin_and_csrf(client, db_session)
    headers = {"x-csrftoken": csrf_token}

    create_payload = {
        "username": "operator_1",
        "password": "operator_password_123",
        "role": "operator",
    }
    response = await client.post("/api/users", json=create_payload, headers=headers)
    assert response.status_code == 201
    user = response.json()
    operator_id = user["id"]
    assert user["username"] == "operator_1"
    assert user["role"] == "operator"

    response = await client.get("/api/users")
    assert response.status_code == 200
    assert len(response.json()) == 2

    response = await client.patch(
        f"/api/users/{operator_id}",
        json={"username": "operator_renamed", "is_active": False},
        headers=headers,
    )
    assert response.status_code == 200
    assert response.json()["username"] == "operator_renamed"
    assert response.json()["is_active"] is False

    operator_result = await db_session.execute(
        AdminUser.__table__.select().where(AdminUser.id == uuid.UUID(operator_id))
    )
    operator_row = operator_result.mappings().first()
    assert operator_row is not None
    assert operator_row["username"] == "operator_renamed"

    client.cookies.clear()
    operator = AdminUser(
        username="operator_access",
        password_hash=hash_password("operator_access_123"),
        role=AdminRole.OPERATOR,
        is_active=True,
    )
    db_session.add(operator)
    await db_session.commit()
    client.cookies.set("access_token", create_access_token(operator.id))
    response = await client.get("/api/users")
    assert response.status_code == 403

    client.cookies.clear()
    client.cookies.set("access_token", create_access_token(admin.id))
    response = await client.get("/setup")
    csrf_token = response.cookies.get("csrftoken") or csrf_token

    response = await client.post(
        f"/api/users/{operator.id}/change-password",
        json={
            "current_password": "operator_access_123",
            "new_password": "new_operator_password_456",
        },
        headers={"x-csrftoken": csrf_token},
    )
    assert response.status_code == 200
    assert response.json()["status"] == "success"

    updated_operator_result = await db_session.execute(
        AdminUser.__table__.select().where(AdminUser.id == operator.id)
    )
    updated_operator_row = updated_operator_result.mappings().first()
    assert updated_operator_row is not None
    assert verify_password("new_operator_password_456", updated_operator_row["password_hash"])

    response = await client.delete(f"/api/users/{operator_id}", headers={"x-csrftoken": csrf_token})
    assert response.status_code == 204


async def test_poll_report_xlsx_generation(client: AsyncClient, db_session: AsyncSession) -> None:
    admin, csrf_token = await _create_admin_and_csrf(client, db_session)
    headers = {"x-csrftoken": csrf_token}

    response = await client.post("/api/polls", json={"title": "Reporte Test"}, headers=headers)
    assert response.status_code == 201
    poll_id = response.json()["id"]

    response = await client.post(
        f"/api/polls/{poll_id}/voter-groups",
        json={"name": "Grupo 1", "weight": 1.0},
        headers=headers,
    )
    assert response.status_code == 201

    response = await client.post(
        f"/api/polls/{poll_id}/categories",
        json={"name": "Categoría A", "order": 0},
        headers=headers,
    )
    assert response.status_code == 201
    category_id = response.json()["id"]

    db_session.add(
        Option(
            poll_id=uuid.UUID(poll_id),
            category_id=uuid.UUID(category_id),
            name="Opción 1",
            order=0,
        )
    )
    await db_session.commit()

    response = await client.post(f"/api/polls/{poll_id}/status", json={"status": "open"}, headers=headers)
    assert response.status_code == 200
    response = await client.post(f"/api/polls/{poll_id}/status", json={"status": "closed"}, headers=headers)
    assert response.status_code == 200
    assert response.json()["status"] == "closed"

    response = await client.get(f"/api/polls/{poll_id}/report.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment;" in response.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["Resumen", "Resultados"]
    summary_ws = workbook["Resumen"]
    results_ws = workbook["Resultados"]
    assert summary_ws["A1"].value == "Reporte de votación"
    assert summary_ws["B4"].value == "Reporte Test"
    assert results_ws["A1"].value == "Categoría"
    assert results_ws["C2"].value == "Opción 1"
