from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.poll import Poll
from app.services.scoring import calculate_poll_results


async def generate_poll_report_xlsx(
    db: AsyncSession,
    poll_id: uuid.UUID,
) -> tuple[bytes, str]:
    """
    Genera un reporte XLSX del poll con una hoja resumen y una hoja de resultados.
    Retorna el contenido binario y un nombre de archivo sugerido.
    """
    poll_result = await calculate_poll_results(db, poll_id)
    if not poll_result:
        raise ValueError("Poll no encontrado")

    poll_query = await db.execute(select(Poll).where(Poll.id == poll_id))
    poll_obj = poll_query.scalar_one_or_none()
    if not poll_obj:
        raise ValueError("Poll no encontrado")

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Resumen"
    results_ws = wb.create_sheet("Resultados")

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9EAF7")

    summary_ws["A1"] = "Reporte de votación"
    summary_ws["A1"].font = title_font
    summary_rows = [
        ("ID del poll", str(poll_id)),
        ("Título", poll_obj.title),
        ("Estado", poll_obj.status.value if hasattr(poll_obj.status, "value") else str(poll_obj.status)),
        ("Generado en", datetime.now(timezone.utc).isoformat()),
    ]

    row_idx = 3
    for label, value in summary_rows:
        summary_ws[f"A{row_idx}"] = label
        summary_ws[f"B{row_idx}"] = value
        summary_ws[f"A{row_idx}"].font = header_font
        row_idx += 1

    summary_ws[f"A{row_idx + 1}"] = "Categorías"
    summary_ws[f"A{row_idx + 1}"].font = header_font

    cat_row = row_idx + 2
    summary_ws[f"A{cat_row}"] = "Categoría"
    summary_ws[f"B{cat_row}"] = "Votos totales"
    summary_ws[f"A{cat_row}"].font = header_font
    summary_ws[f"B{cat_row}"].font = header_font
    summary_ws[f"A{cat_row}"].fill = fill
    summary_ws[f"B{cat_row}"].fill = fill

    for cat in poll_result["categories"]:
        cat_row += 1
        summary_ws[f"A{cat_row}"] = cat["category_name"]
        summary_ws[f"B{cat_row}"] = cat["total_votes"]

    results_headers = [
        "Categoría",
        "Posición",
        "Opción",
        "Score",
        "Votos totales",
        "Foto",
    ]
    for col_idx, header in enumerate(results_headers, start=1):
        cell = results_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill

    current_row = 2
    for category in poll_result["categories"]:
        for position, result in enumerate(category["results"], start=1):
            results_ws.cell(row=current_row, column=1, value=category["category_name"])
            results_ws.cell(row=current_row, column=2, value=position)
            results_ws.cell(row=current_row, column=3, value=result["option_name"])
            results_ws.cell(row=current_row, column=4, value=result["score"])
            results_ws.cell(row=current_row, column=5, value=result["total_votes"])
            results_ws.cell(row=current_row, column=6, value=result["photo_url"])
            current_row += 1

    for ws in (summary_ws, results_ws):
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    filename = f"poll-{poll_id}.xlsx"
    return buffer.getvalue(), filename
